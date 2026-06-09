import json
import requests
import pandas as pd

from pathlib import Path
from urllib.parse import quote
from openpyxl.styles import Font, PatternFill, Alignment


# ============================================================
# 1. 사용자 설정
# ============================================================

SERVICE_KEY = ""

START_DATE = "20080101"  # yyyymmdd
END_DATE = "20081231"    # yyyymmdd

OUTPUT_DIR = Path("./smartfarm_hanwoo_output")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

OUTPUT_XLSX = OUTPUT_DIR / f"smartfarm_hanwoo_grade_{START_DATE}_{END_DATE}.xlsx"
OUTPUT_CSV = OUTPUT_DIR / f"smartfarm_hanwoo_grade_{START_DATE}_{END_DATE}.csv"
RAW_JSON = OUTPUT_DIR / f"smartfarm_hanwoo_grade_raw_{START_DATE}_{END_DATE}.json"


# ============================================================
# 2. API 요청
# ============================================================

BASE_URL = (
    "http://www.smartfarmkorea.net/Agree_WS/webservices/"
    "StockRestService/getShipmntDataList"
)

encoded_key = quote(SERVICE_KEY, safe="")

url = f"{BASE_URL}/{encoded_key}/{START_DATE}/{END_DATE}"

print("요청 URL:")
print(url.replace(encoded_key, "SERVICE_KEY_HIDDEN"))

res = requests.get(
    url,
    headers={
        "Accept": "application/json",
        "User-Agent": "Mozilla/5.0"
    },
    timeout=300
)

print("status_code:", res.status_code)
print("content-type:", res.headers.get("content-type"))

if res.status_code != 200:
    print(res.text[:2000])
    raise RuntimeError("API 요청 실패")

try:
    data = res.json()
except Exception:
    print(res.text[:3000])
    raise RuntimeError("JSON 파싱 실패")

if isinstance(data, dict):
    data = [data]

if not isinstance(data, list):
    raise RuntimeError(f"예상하지 못한 응답 타입: {type(data)}")

print(f"수집 레코드 수: {len(data):,}")

# 원본 JSON 저장
with RAW_JSON.open("w", encoding="utf-8") as f:
    json.dump(data, f, ensure_ascii=False, indent=2)


# ============================================================
# 3. DataFrame 변환
# ============================================================

df = pd.json_normalize(data)


# ============================================================
# 4. ID 컬럼 문자열 보존
# ============================================================

id_cols = [
    "farmId",
    "farmManageNo",
    "indvdNo",
    "histNo",
    "pedigreeNo"
]

for col in id_cols:
    if col in df.columns:
        # 결측값은 그대로 두고, 값이 있는 경우 문자열로 고정
        df[col] = df[col].apply(lambda x: "" if pd.isna(x) else str(x).strip())


# ============================================================
# 5. 날짜 컬럼 문자열 보존 + 날짜형 파생
# ============================================================

date_cols = [
    "shipmntDe",
    "slauDe",
    "gradJdgmntDe",
    "birthDe"
]

for col in date_cols:
    if col in df.columns:
        df[col] = df[col].apply(lambda x: "" if pd.isna(x) else str(x).strip())
        df[col + "_dt"] = pd.to_datetime(df[col], format="%Y%m%d", errors="coerce")


# ============================================================
# 6. 숫자 컬럼 변환
# ============================================================

numeric_cols = [
    "carcassWt",          # 도체중량
    "backFatThick",       # 등지방두께
    "loinEyeAr",          # 등심단면적
    "meatQuantityIdex",   # 육량지수
    "marbling",           # 근내지방도
    "meatColor",          # 육색
    "fatColor",           # 지방색
    "contextDgree",       # 조직감
    "mtrdg",              # 성숙도
    "auctionCost"         # 경락가격
]

for col in numeric_cols:
    if col in df.columns:
        df[col] = pd.to_numeric(df[col], errors="coerce")


# ============================================================
# 7. 컬럼명 한글 설명 추가용 매핑
# ============================================================

column_desc = {
    "farmId": "농장 ID",
    "indvdNo": "개체번호",
    "histNo": "이력번호",
    "shipmntDe": "출하 일자",
    "slauIpla": "도축장명",
    "slauDe": "도축 일자",
    "meatQuantityGrade": "육량 등급 코드",
    "meatQualityGrade": "육질 등급 코드",
    "carcassWt": "도체중량",
    "backFatThick": "등지방두께",
    "auctionCost": "경락가격",
    "gradJdgmntDe": "등급판정일자",
    "meatQuantityIdex": "육량지수",
    "loinEyeAr": "등심단면적",
    "gradRevisn": "등급 보정",
    "marbling": "근내지방도",
    "meatColor": "육색",
    "fatColor": "지방색",
    "contextDgree": "조직감",
    "mtrdg": "성숙도",
    "lastGrad": "최종등급",
    "statusMessage": "상태 메시지",
    "statusCode": "상태 코드",
}


# ============================================================
# 8. 컬럼 순서 정리
# ============================================================

preferred_order = [
    "farmId",
    "indvdNo",
    "histNo",
    "shipmntDe",
    "slauIpla",
    "slauDe",
    "gradJdgmntDe",
    "carcassWt",
    "backFatThick",
    "loinEyeAr",
    "meatQuantityGrade",
    "meatQualityGrade",
    "meatQuantityIdex",
    "marbling",
    "meatColor",
    "fatColor",
    "contextDgree",
    "mtrdg",
    "gradRevisn",
    "lastGrad",
    "auctionCost",
    "statusCode",
    "statusMessage",
]

existing = [c for c in preferred_order if c in df.columns]
remaining = [c for c in df.columns if c not in existing]
df = df[existing + remaining]


# ============================================================
# 9. 요약 테이블 생성
# ============================================================

summary_by_date = None
if "shipmntDe" in df.columns:
    summary_by_date = (
        df.groupby("shipmntDe", dropna=False)
        .agg(
            n=("histNo", "count"),
            avg_carcassWt=("carcassWt", "mean"),
            avg_backFatThick=("backFatThick", "mean"),
            avg_loinEyeAr=("loinEyeAr", "mean"),
            avg_marbling=("marbling", "mean"),
            avg_meatColor=("meatColor", "mean"),
            avg_fatColor=("fatColor", "mean"),
            avg_contextDgree=("contextDgree", "mean"),
            avg_mtrdg=("mtrdg", "mean"),
            avg_auctionCost=("auctionCost", "mean"),
        )
        .reset_index()
    )

summary_by_farm = None
if "farmId" in df.columns:
    summary_by_farm = (
        df.groupby("farmId", dropna=False)
        .agg(
            n=("histNo", "count"),
            avg_carcassWt=("carcassWt", "mean"),
            avg_backFatThick=("backFatThick", "mean"),
            avg_loinEyeAr=("loinEyeAr", "mean"),
            avg_marbling=("marbling", "mean"),
            avg_meatColor=("meatColor", "mean"),
            avg_fatColor=("fatColor", "mean"),
            avg_contextDgree=("contextDgree", "mean"),
            avg_mtrdg=("mtrdg", "mean"),
            avg_auctionCost=("auctionCost", "mean"),
        )
        .reset_index()
        .sort_values("n", ascending=False)
    )

codebook = pd.DataFrame(
    [{"column": col, "description": column_desc.get(col, "")} for col in df.columns]
)


# ============================================================
# 10. CSV 저장
# ============================================================

# CSV는 Excel로 바로 열면 앞자리 0이 사라질 수 있음.
# 그래도 원본 보관용으로 저장.
df.to_csv(OUTPUT_CSV, index=False, encoding="utf-8-sig")


# ============================================================
# 11. XLSX 저장: ID 컬럼 텍스트 서식 강제
# ============================================================

with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="raw_data", index=False)
    codebook.to_excel(writer, sheet_name="codebook", index=False)

    if summary_by_date is not None:
        summary_by_date.to_excel(writer, sheet_name="summary_by_date", index=False)

    if summary_by_farm is not None:
        summary_by_farm.to_excel(writer, sheet_name="summary_by_farm", index=False)

    workbook = writer.book

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"

        if sheet.max_row >= 2:
            sheet.auto_filter.ref = sheet.dimensions

        # 헤더 스타일
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
            cell.alignment = Alignment(horizontal="center")

        # 컬럼 폭 조정
        for col_cells in sheet.columns:
            max_len = 0
            col_letter = col_cells[0].column_letter

            for cell in col_cells:
                if cell.value is None:
                    continue
                max_len = max(max_len, len(str(cell.value)))

            sheet.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 35)

    # raw_data 시트의 ID 컬럼을 텍스트 서식으로 강제
    ws = workbook["raw_data"]

    header_to_col_idx = {
        cell.value: cell.column
        for cell in ws[1]
        if cell.value is not None
    }

    text_cols = [
        "farmId",
        "farmManageNo",
        "indvdNo",
        "histNo",
        "pedigreeNo",
        "shipmntDe",
        "slauDe",
        "gradJdgmntDe",
    ]

    for col_name in text_cols:
        if col_name not in header_to_col_idx:
            continue

        col_idx = header_to_col_idx[col_name]

        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            if cell.value is not None:
                cell.value = str(cell.value)
                cell.number_format = "@"


print("저장 완료")
print(f"Excel: {OUTPUT_XLSX.resolve()}")
print(f"CSV:   {OUTPUT_CSV.resolve()}")
print(f"JSON:  {RAW_JSON.resolve()}")
print()
print(df.head())
print(df.dtypes)